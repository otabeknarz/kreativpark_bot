import openpyxl
import datetime


def make_stats(people, purposes, days):
    workbook = openpyxl.Workbook()

    people_sheet = workbook.create_sheet("Odamlar ro'yxati", index=0)
    other_stats = workbook.create_sheet("Boshqa statistikalar", index=1)

    other_stats.cell(row=1, column=1).value = "№"
    other_stats.cell(row=1, column=2).value = "Qayer"
    other_stats.cell(row=1, column=3).value = "Kirganlar soni"

    other_stats.column_dimensions["A"].width = 3
    other_stats.column_dimensions["B"].width = 15
    other_stats.column_dimensions["C"].width = 10

    row = 1
    for name, count in purposes.items():
        other_stats.cell(row=row + 1, column=1).value = str(row)
        other_stats.cell(row=row + 1, column=2).value = name
        other_stats.cell(row=row + 1, column=3).value = count
        row += 1

    if "Sheet" in workbook.sheetnames:
        del workbook["Sheet"]

    people_sheet.cell(row=1, column=1).value = "№"
    people_sheet.cell(row=1, column=2).value = "Ism Familiya"
    people_sheet.cell(row=1, column=3).value = "Telegram ID"
    people_sheet.cell(row=1, column=4).value = "Telefon raqami"
    people_sheet.cell(row=1, column=5).value = "Passport Ma'lumot"
    people_sheet.cell(row=1, column=6).value = "Registratsiya sanasi"
    row = 1
    for p in people:
        people_sheet.cell(row=row + 1, column=1).value = str(row)
        people_sheet.cell(row=row + 1, column=2).value = p["name"]
        people_sheet.cell(row=row + 1, column=3).value = p["ID"]
        people_sheet.cell(row=row + 1, column=4).value = p["phone_number"]
        people_sheet.cell(row=row + 1, column=5).value = p["passport_data"]
        created_at = datetime.datetime.fromisoformat(p["created_at"])
        people_sheet.cell(row=row + 1, column=6).value = created_at.strftime(
            "%d-%m-%Y | %H:%M"
        )
        row += 1
    name = (
        "xlsx/" + datetime.datetime.today().strftime("%d-%m-%Y__" + str(days)) + ".xlsx"
    )
    people_sheet.column_dimensions["A"].width = 3
    people_sheet.column_dimensions["B"].width = 30
    people_sheet.column_dimensions["C"].width = 15
    people_sheet.column_dimensions["D"].width = 20
    people_sheet.column_dimensions["E"].width = 20
    people_sheet.column_dimensions["F"].width = 20
    workbook.save(name)
    return name
